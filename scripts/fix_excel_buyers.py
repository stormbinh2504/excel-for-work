#!/usr/bin/env python3
"""
Fix buyer MST and name in an Excel file by propagating values from parent rows to child rows.

Logic:
- Parent rows: rows where column A is empty and column B ("Mẫu số") is non-empty.
- Child rows: rows where column A contains a number (1,2,3...).
- For each child row, if column F (MST người mua) or column G (Tên người mua) is empty or whitespace,
  copy the corresponding value from the most recent parent row above it.

Notes:
- Headers on rows 9 and 10 are preserved and skipped.
- By default this script writes a new file with suffix "_fixed.xlsx"; use --overwrite to replace input.

Requires: openpyxl (pip install openpyxl)

Usage:
  python scripts/fix_excel_buyers.py path/to/DOANHTHUTHEOSANPHAM-02.2026-01.04.2026.xlsx

"""
from __future__ import annotations

import argparse
import os
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook


def is_child_row(cell_value) -> bool:
    """Return True if column A cell indicates a child row (numeric STT)."""
    if cell_value is None:
        return False
    if isinstance(cell_value, (int, float)):
        return True
    try:
        s = str(cell_value).strip()
    except Exception:
        return False
    if s == "":
        return False
    # treat purely numeric strings as STT
    if s.isdigit():
        return True
    # some STT might be like '1.' or '1)'
    if s.rstrip('. )').isdigit():
        return True
    return False


def is_nonempty(val) -> bool:
    return val is not None and str(val).strip() != ""


def fix_file(input_path: str, output_path: str, header_rows=(9, 10)) -> dict:
    wb = load_workbook(input_path)
    ws = wb.active

    current_parent_F: Optional[object] = None
    current_parent_G: Optional[object] = None

    changed_count = 0
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))

    for idx, row in enumerate(rows, start=1):
        if idx in header_rows:
            # skip header rows entirely
            continue

        cellA = ws.cell(row=idx, column=1)
        cellB = ws.cell(row=idx, column=2)
        cellF = ws.cell(row=idx, column=6)
        cellG = ws.cell(row=idx, column=7)

        if not is_child_row(cellA.value):
            # Potential parent row if column B has content
            if is_nonempty(cellB.value):
                # update current parent values from this row
                current_parent_F = cellF.value
                current_parent_G = cellG.value
            # else: neither parent nor child, ignore
        else:
            # child row: fill missing/incomplete values from current parent if available
            if current_parent_F is not None and (cellF.value is None or str(cellF.value).strip() == ""):
                cellF.value = current_parent_F
                changed_count += 1
            if current_parent_G is not None and (cellG.value is None or str(cellG.value).strip() == ""):
                cellG.value = current_parent_G
                changed_count += 1

    # Save workbook
    wb.save(output_path)
    return {"changed_cells": changed_count, "output": output_path}


def main() -> None:
    p = argparse.ArgumentParser(description="Propagate buyer MST and name from parent rows to child rows in an Excel file.")
    p.add_argument("input", help="Input .xlsx file path")
    p.add_argument("-o", "--output", help="Output file path (default: <input>_fixed.xlsx)")
    p.add_argument("--overwrite", action="store_true", help="Overwrite input file (will make a timestamped backup)")
    args = p.parse_args()

    input_path = args.input
    if not os.path.isfile(input_path):
        raise SystemExit(f"Input file not found: {input_path}")

    if args.overwrite:
        # create backup
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = f"{input_path}.backup.{ts}"
        os.rename(input_path, backup)
        output_path = input_path
        print(f"Created backup: {backup}")
    else:
        if args.output:
            output_path = args.output
        else:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_fixed{ext}"

    print(f"Processing: {input_path} -> {output_path}")
    result = fix_file(input_path, output_path)
    print(f"Done. Changed cells: {result['changed_cells']}")


if __name__ == "__main__":
    main()
